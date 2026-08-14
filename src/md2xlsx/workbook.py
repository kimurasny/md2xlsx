"""中間ブロックを openpyxl のワークシートへ描画する。"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.drawing.image import Image as XlsxImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .blocks import (
    Block,
    CodeBlock,
    Heading,
    HorizontalRule,
    ImageBlock,
    ListBlock,
    Paragraph,
    Run,
    Section,
    Table,
)
from .image_handler import ImageError, is_data_uri, is_remote, resolve_image
from .utils import unique_sheet_name

# 日本語文書での可読性を優先し、既定の本文フォントはメイリオとする。
BODY_FONT_NAME = "Meiryo"
CODE_FONT_NAME = "Consolas"
BODY_FONT_SIZE = 11

TEXT_COLUMN_WIDTH = 48
TABLE_COLUMN_WIDTH = 20
INDENT_PREFIX = "    "

# 見出しレベルごとのフォントサイズ。
_HEADING_SIZES = {1: 16, 2: 14, 3: 12.5, 4: 12, 5: 11, 6: 11}

_HEADING_FILL = PatternFill("solid", fgColor="EFEFEF")
_TABLE_HEADER_FILL = PatternFill("solid", fgColor="DDEBF7")
_CODE_FILL = PatternFill("solid", fgColor="F5F5F5")

_THIN_SIDE = Side(style="thin", color="B0B0B0")
_CELL_BORDER = Border(left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE)

_WRAP_TOP = Alignment(wrap_text=True, vertical="top")
_NOWRAP_TOP = Alignment(wrap_text=False, vertical="top")

# 1 ピクセルあたりの行高（ポイント）。
_POINTS_PER_PIXEL = 0.75


class SheetWriter:
    """1 シート分の描画状態（現在行・列幅・警告）を保持する。"""

    def __init__(self, sheet: Worksheet, base_dir: Path) -> None:
        self._sheet = sheet
        self._base_dir = base_dir
        self._row = 1
        self._max_table_columns = 1
        self.warnings: list[str] = []
        self.image_count = 0

    def write_blocks(self, blocks: list[Block]) -> None:
        previous: Block | None = None
        for block in blocks:
            if previous is not None and self._needs_spacer(previous, block):
                self._row += 1
            self._write_block(block)
            previous = block
        self._apply_column_widths()

    @staticmethod
    def _needs_spacer(previous: Block, current: Block) -> bool:
        """読みやすさのために空行を挟むかどうかを判定する。"""
        if isinstance(previous, ListBlock) and isinstance(current, ListBlock):
            return False
        return True

    def _write_block(self, block: Block) -> None:
        if isinstance(block, Heading):
            self._write_heading(block)
        elif isinstance(block, Paragraph):
            self._write_paragraph(block)
        elif isinstance(block, ListBlock):
            self._write_list(block)
        elif isinstance(block, Table):
            self._write_table(block)
        elif isinstance(block, CodeBlock):
            self._write_code(block)
        elif isinstance(block, ImageBlock):
            self._write_image(block)
        elif isinstance(block, HorizontalRule):
            self._write_rule()

    def _cell(self, column: int = 1):
        return self._sheet.cell(row=self._row, column=column)

    def _write_heading(self, heading: Heading) -> None:
        cell = self._cell()
        cell.value = heading.text
        size = _HEADING_SIZES.get(heading.level, 11)
        cell.font = Font(name=BODY_FONT_NAME, size=size, bold=True)
        cell.alignment = _WRAP_TOP
        if heading.level <= 2:
            cell.fill = _HEADING_FILL
        self._row += 1

    def _write_paragraph(self, paragraph: Paragraph) -> None:
        cell = self._cell()
        text = paragraph.text
        if paragraph.quote:
            cell.value = _rich_text(paragraph.runs, prefix="> ")
            cell.font = Font(
                name=BODY_FONT_NAME, size=BODY_FONT_SIZE, italic=True, color="595959"
            )
        else:
            cell.value = _rich_text(paragraph.runs)
            cell.font = Font(name=BODY_FONT_NAME, size=BODY_FONT_SIZE)
        cell.alignment = _WRAP_TOP
        if not text.strip():
            cell.value = None
        self._row += 1

    def _write_list(self, block: ListBlock) -> None:
        for item in block.items:
            cell = self._cell()
            prefix = INDENT_PREFIX * item.indent
            marker = f"{item.marker} " if item.marker else "  "
            cell.value = _rich_text(item.runs, prefix=f"{prefix}{marker}")
            cell.font = Font(name=BODY_FONT_NAME, size=BODY_FONT_SIZE)
            cell.alignment = _WRAP_TOP
            self._row += 1

    def _write_table(self, table: Table) -> None:
        column_count = table.column_count
        if column_count == 0:
            return
        self._max_table_columns = max(self._max_table_columns, column_count)

        if table.header:
            self._write_table_row(table.header, column_count, header=True)
        for row in table.rows:
            self._write_table_row(row, column_count, header=False)

    def _write_table_row(self, values: list[str], column_count: int, header: bool) -> None:
        for index in range(column_count):
            cell = self._cell(column=index + 1)
            cell.value = values[index] if index < len(values) else None
            cell.alignment = _WRAP_TOP
            cell.border = _CELL_BORDER
            cell.font = Font(name=BODY_FONT_NAME, size=BODY_FONT_SIZE, bold=header)
            if header:
                cell.fill = _TABLE_HEADER_FILL
        self._row += 1

    def _write_code(self, block: CodeBlock) -> None:
        label = f"[code: {block.language}]" if block.language else "[code]"
        label_cell = self._cell()
        label_cell.value = label
        label_cell.font = Font(name=CODE_FONT_NAME, size=9, color="808080")
        self._row += 1
        for line in block.lines:
            cell = self._cell()
            cell.value = line
            cell.font = Font(name=CODE_FONT_NAME, size=10)
            cell.fill = _CODE_FILL
            cell.alignment = _NOWRAP_TOP
            self._row += 1

    def _write_rule(self) -> None:
        cell = self._cell()
        cell.value = "―" * 20
        cell.font = Font(name=BODY_FONT_NAME, size=BODY_FONT_SIZE, color="A6A6A6")
        self._row += 1

    def _write_image(self, block: ImageBlock) -> None:
        if is_remote(block.src) or is_data_uri(block.src):
            self._write_note(f"[External image (not downloaded): {block.src}]")
            self.warnings.append(f"外部画像は取り込みませんでした: {block.src}")
            return
        try:
            resolved = resolve_image(block.src, self._base_dir)
            image = XlsxImage(resolved.path)
            image.width = resolved.width
            image.height = resolved.height
            anchor = f"{get_column_letter(1)}{self._row}"
            self._sheet.add_image(image, anchor)
        except (ImageError, OSError, ValueError) as error:
            self._write_note(f"[Image unavailable: {block.src}]")
            self.warnings.append(f"画像を埋め込めませんでした: {block.src} ({error})")
            return

        # 画像と後続コンテンツが重ならないよう、アンカー行の高さを画像に合わせる。
        self._sheet.row_dimensions[self._row].height = resolved.height * _POINTS_PER_PIXEL
        self._row += 1
        if block.alt.strip():
            caption = self._cell()
            caption.value = f"図: {block.alt.strip()}"
            caption.font = Font(name=BODY_FONT_NAME, size=9, italic=True, color="595959")
            caption.alignment = _WRAP_TOP
            self._row += 1
        self.image_count += 1

    def _write_note(self, text: str) -> None:
        cell = self._cell()
        cell.value = text
        cell.font = Font(name=BODY_FONT_NAME, size=BODY_FONT_SIZE, italic=True, color="C00000")
        cell.alignment = _WRAP_TOP
        self._row += 1

    def _apply_column_widths(self) -> None:
        self._sheet.column_dimensions["A"].width = TEXT_COLUMN_WIDTH
        for index in range(2, self._max_table_columns + 1):
            letter = get_column_letter(index)
            self._sheet.column_dimensions[letter].width = TABLE_COLUMN_WIDTH


def _rich_text(runs: list[Run], prefix: str = "") -> CellRichText | str | None:
    """装飾付き文字列を Excel のリッチテキストへ変換する。

    装飾が無い場合は通常の文字列を返す（不要なリッチテキスト化を避ける）。
    """
    pieces: list[str | TextBlock] = []
    if prefix:
        pieces.append(prefix)
    decorated = False
    for run in runs:
        if not run.text:
            continue
        if run.bold or run.italic or run.code:
            decorated = True
            font = InlineFont(
                b=run.bold,
                i=run.italic,
                rFont=CODE_FONT_NAME if run.code else BODY_FONT_NAME,
                sz=BODY_FONT_SIZE,
            )
            pieces.append(TextBlock(font, run.text))
        else:
            pieces.append(run.text)
    if not pieces:
        return None
    if not decorated:
        return "".join(str(piece) for piece in pieces)
    return CellRichText(pieces)


def apply_default_font(
    workbook: Workbook,
    name: str = BODY_FONT_NAME,
    size: float = BODY_FONT_SIZE,
) -> None:
    """Workbook 全体の既定フォントを差し替える。

    openpyxl には既定フォントを設定する API がないため、Normal スタイルが
    参照するフォントテーブル先頭の定義を置き換える。これにより、このツールが
    値を書き込んでいないセル（利用者が後から入力するセル）も同じフォントになる。
    """
    workbook._fonts[0] = Font(name=name, size=size)


class WorkbookBuilder:
    """セクション列から Workbook を組み立てる。"""

    def __init__(self, base_dir: Path) -> None:
        self._base_dir = base_dir
        self.warnings: list[str] = []
        self.image_count = 0

    def build(self, sections: list[Section]) -> Workbook:
        workbook = Workbook()
        apply_default_font(workbook)
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        used_names: set[str] = set()
        if not sections:
            sections = [Section(title="Introduction", is_intro=True)]

        for section in sections:
            name = unique_sheet_name(section.title, used_names)
            sheet = workbook.create_sheet(title=name)
            writer = SheetWriter(sheet, self._base_dir)
            writer.write_blocks(section.blocks)
            self.warnings.extend(writer.warnings)
            self.image_count += writer.image_count
        return workbook
