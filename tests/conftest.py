"""テスト共通のヘルパー。"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from PIL import Image as PilImage


@pytest.fixture
def write_md(tmp_path: Path):
    """UTF-8 で Markdown ファイルを書き出すヘルパーを返す。"""

    def _write(relative: str, text: str) -> Path:
        path = tmp_path / relative
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(text, encoding="utf-8")
        return path

    return _write


@pytest.fixture
def make_png():
    """指定サイズの PNG 画像を作成するヘルパーを返す。"""

    def _make(path: Path, size: tuple[int, int] = (40, 30)) -> Path:
        path.parent.mkdir(parents=True, exist_ok=True)
        PilImage.new("RGB", size, color=(12, 120, 200)).save(path)
        return path

    return _make


def open_workbook(path: Path) -> Workbook:
    return load_workbook(path)


def column_values(sheet, column: str = "A") -> list[str]:
    """指定列の値を、空セルを除いた文字列リストで返す。"""
    values: list[str] = []
    for row in range(1, sheet.max_row + 1):
        value = sheet[f"{column}{row}"].value
        if value is None:
            continue
        values.append(str(value))
    return values


def row_values(sheet, row: int, width: int) -> list[str | None]:
    return [sheet.cell(row=row, column=index + 1).value for index in range(width)]


# EMU（English Metric Unit）からピクセルへの換算係数。
_EMU_PER_PIXEL = 9525


def image_size_px(image) -> tuple[int, int]:
    """XLSX に保存された描画サイズ（ピクセル）を返す。"""
    ext = image.anchor.ext
    return round(ext.cx / _EMU_PER_PIXEL), round(ext.cy / _EMU_PER_PIXEL)
