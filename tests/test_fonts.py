"""既定フォント（メイリオ）が各セルへ適用されることのテスト。"""

from __future__ import annotations

from pathlib import Path

from conftest import open_workbook
from openpyxl import load_workbook

from md2xlsx.converter import convert_file
from md2xlsx.workbook import BODY_FONT_NAME, CODE_FONT_NAME

MARKDOWN = """## 概要

これは**太字**を含む本文です。

- 項目A

1. 一番目

> 引用文

| Name | Type |
|------|------|
| id | string |

```python
value = 1
```

![なし](./missing.png)
"""


def _cell_font_names(sheet) -> set[str]:
    names: set[str] = set()
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            names.add(cell.font.name)
    return names


def test_body_font_is_meiryo(write_md, tmp_path: Path) -> None:
    assert BODY_FONT_NAME == "Meiryo"

    source = write_md("a.md", MARKDOWN)
    output = tmp_path / "a.xlsx"
    convert_file(source, output)

    sheet = open_workbook(output)["概要"]
    # 本文・見出し・リスト・引用・テーブル・注記はメイリオ、コードのみ等幅フォント。
    assert _cell_font_names(sheet) == {BODY_FONT_NAME, CODE_FONT_NAME}


def test_untouched_cell_uses_meiryo(write_md, tmp_path: Path) -> None:
    source = write_md("a.md", "## 概要\n\n本文\n")
    output = tmp_path / "a.xlsx"
    convert_file(source, output)

    sheet = open_workbook(output)["概要"]
    assert sheet["Z100"].font.name == BODY_FONT_NAME


def test_rich_text_runs_use_meiryo(write_md, tmp_path: Path) -> None:
    source = write_md("a.md", "## 概要\n\nこれは**太字**と`コード`です。\n")
    output = tmp_path / "a.xlsx"
    convert_file(source, output)

    sheet = load_workbook(output, rich_text=True)["概要"]
    value = next(
        cell.value
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None and "太字" in str(cell.value)
    )
    fonts = {block.font.rFont for block in value if hasattr(block, "font")}
    assert fonts == {BODY_FONT_NAME, CODE_FONT_NAME}
